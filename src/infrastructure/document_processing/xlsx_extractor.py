"""XLSX extractor using openpyxl for Excel file parsing."""

import asyncio
import time
from pathlib import Path
from typing import Any

import pandas as pd
import structlog
from openpyxl import load_workbook

logger = structlog.get_logger(__name__)


class XLSXExtractor:
    """Excel file extractor using openpyxl.
    
    Supports specific sheet selection, handles formulas and merged cells.
    Returns structured data as DataFrames.
    """

    def __init__(self, data_only: bool = True):
        """Initialize XLSX extractor.
        
        Args:
            data_only: If True, read cached values instead of formulas (default: True)
        """
        self.data_only = data_only
        self.logger = logger.bind(component="xlsx_extractor")

    async def extract(
        self,
        file_path: str | Path,
        sheet_name: str | int | None = None,
        header_row: int = 1,
    ) -> dict[str, pd.DataFrame]:
        """Extract data from Excel file.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name or index (0-based). None = all sheets
            header_row: Row number containing headers (1-based, default: 1)
            
        Returns:
            Dictionary mapping sheet names to DataFrames
            
        Example:
            >>> extractor = XLSXExtractor()
            >>> sheets = await extractor.extract("assets.xlsx", sheet_name="Assets")
            >>> df = sheets["Assets"]
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        start_time = time.monotonic()
        self.logger.info("xlsx_extract_start", file_path=str(file_path), sheet_name=sheet_name)

        try:
            result = await asyncio.to_thread(
                self._extract_sync, file_path, sheet_name, header_row
            )
            duration_ms = (time.monotonic() - start_time) * 1000
            self.logger.info(
                "xlsx_extract_complete",
                file_path=str(file_path),
                duration_ms=round(duration_ms, 2),
                sheet_count=len(result),
                sheets=list(result.keys()),
            )
            return result
        except Exception as e:
            duration_ms = (time.monotonic() - start_time) * 1000
            self.logger.exception(
                "xlsx_extract_failed",
                error=str(e),
                file_path=str(file_path),
                duration_ms=round(duration_ms, 2),
            )
            raise

    def _extract_sync(
        self,
        file_path: Path,
        sheet_name: str | int | None,
        header_row: int,
    ) -> dict[str, pd.DataFrame]:
        """Synchronous XLSX extraction (runs in thread pool)."""
        wb = load_workbook(str(file_path), data_only=self.data_only, read_only=False)
        
        try:
            result: dict[str, pd.DataFrame] = {}
            
            # Determine which sheets to extract
            if sheet_name is None:
                sheets_to_extract = wb.sheetnames
            elif isinstance(sheet_name, int):
                if sheet_name < 0 or sheet_name >= len(wb.sheetnames):
                    raise ValueError(f"Sheet index {sheet_name} out of range")
                sheets_to_extract = [wb.sheetnames[sheet_name]]
            else:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                sheets_to_extract = [sheet_name]
            
            # Extract each sheet
            for sheet in sheets_to_extract:
                ws = wb[sheet]
                self.logger.debug("xlsx_processing_sheet", file_path=str(file_path), sheet=sheet)

                # Read all data
                data: list[list[Any]] = []
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
                
                if not data:
                    result[sheet] = pd.DataFrame()
                    continue
                
                # Convert to DataFrame with header
                if header_row > 0 and header_row <= len(data):
                    headers = data[header_row - 1]
                    body = data[header_row:]
                    df = pd.DataFrame(body, columns=headers)
                else:
                    df = pd.DataFrame(data)
                
                # Handle merged cells by forward-filling
                df = df.ffill()
                
                result[sheet] = df
            
            return result
        finally:
            wb.close()

    async def extract_single_sheet(
        self,
        file_path: str | Path,
        sheet_name: str | int = 0,
        header_row: int = 1,
    ) -> pd.DataFrame:
        """Extract data from a single sheet.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index (0-based)
            header_row: Row number containing headers (1-based)
            
        Returns:
            DataFrame with sheet data
        """
        sheets = await self.extract(file_path, sheet_name, header_row)
        sheet_key = list(sheets.keys())[0]
        return sheets[sheet_key]

    async def get_sheet_names(self, file_path: str | Path) -> list[str]:
        """Get list of sheet names in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        file_path = Path(file_path)
        
        async def _get_names():
            wb = load_workbook(str(file_path), read_only=True)
            try:
                return wb.sheetnames
            finally:
                wb.close()
        
        return await asyncio.to_thread(_get_names)
