"""Assets/liabilities XLSX generator using openpyxl."""

import random
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

from src.data_generation.profile import ApplicantProfile


def _generate_asset_categories(profile: ApplicantProfile, rng: random.Random) -> dict[str, Decimal]:
    """Generate individual asset category values within specified ranges."""
    cash_and_deposits = Decimal(str(rng.uniform(1_000, 50_000))).quantize(Decimal("0.01"))
    savings_accounts = Decimal(str(rng.uniform(5_000, 200_000))).quantize(Decimal("0.01"))
    investment_accounts = Decimal(str(rng.uniform(0, 500_000))).quantize(Decimal("0.01"))
    retirement_accounts = Decimal(str(rng.uniform(0, 300_000))).quantize(Decimal("0.01"))

    if profile.housing_status == "owned":
        real_estate_value = Decimal(str(rng.uniform(500_000, 2_000_000))).quantize(Decimal("0.01"))
    else:
        real_estate_value = Decimal("0.00")

    vehicle_value = Decimal(str(rng.uniform(20_000, 200_000))).quantize(Decimal("0.01"))
    other_assets = Decimal(str(rng.uniform(0, 50_000))).quantize(Decimal("0.01"))

    return {
        "cash_and_deposits": cash_and_deposits,
        "savings_accounts": savings_accounts,
        "investment_accounts": investment_accounts,
        "retirement_accounts": retirement_accounts,
        "real_estate_value": real_estate_value,
        "vehicle_value": vehicle_value,
        "other_assets": other_assets,
    }


def _generate_liability_categories(
    profile: ApplicantProfile, credit_data: dict | None, credit_facilities: list[dict] | None, rng: random.Random
) -> dict[str, Decimal]:
    """Generate individual liability category values within specified ranges.

    When credit_data and credit_facilities are provided, coordinates credit-related
    liabilities with the credit report for cross-document consistency.
    """
    if profile.housing_status == "owned":
        mortgage_balance = Decimal(str(rng.uniform(200_000, 1_000_000))).quantize(Decimal("0.01"))
    else:
        mortgage_balance = Decimal("0.00")

    # Coordinate with credit report for cross-document consistency
    if credit_facilities and credit_data:
        # Sum up credit card and personal loan facilities from the credit report
        cc_balances = [
            Decimal(str(f["current_balance"]))
            for f in credit_facilities
            if f["facility_type"] == "credit_card"
        ]
        pl_balances = [
            Decimal(str(f["current_balance"]))
            for f in credit_facilities
            if f["facility_type"] == "personal_loan"
        ]
        credit_card_debt = sum(cc_balances, Decimal("0")).quantize(Decimal("0.01"))
        personal_loans = sum(pl_balances, Decimal("0")).quantize(Decimal("0.01"))
    elif credit_data is not None and "total_outstanding_balance" in credit_data:
        total_credit = Decimal(str(credit_data["total_outstanding_balance"])).quantize(Decimal("0.01"))
        credit_card_debt = (total_credit * Decimal("0.7")).quantize(Decimal("0.01"))
        personal_loans = (total_credit * Decimal("0.3")).quantize(Decimal("0.01"))
    else:
        credit_card_debt = Decimal(str(rng.uniform(0, 50_000))).quantize(Decimal("0.01"))
        personal_loans = Decimal(str(rng.uniform(0, 100_000))).quantize(Decimal("0.01"))

    student_loans = Decimal(str(rng.uniform(0, 50_000))).quantize(Decimal("0.01"))
    other_liabilities = Decimal(str(rng.uniform(0, 20_000))).quantize(Decimal("0.01"))

    return {
        "mortgage_balance": mortgage_balance,
        "personal_loans": personal_loans,
        "credit_card_debt": credit_card_debt,
        "student_loans": student_loans,
        "other_liabilities": other_liabilities,
    }


def _build_asset_details(assets: dict[str, Decimal], rng: random.Random) -> list[dict[str, Any]]:
    """Build detailed asset records for the asset_details JSONB field."""
    details = []
    asset_types = {
        "cash_and_deposits": "Cash & Deposits",
        "savings_accounts": "Savings Accounts",
        "investment_accounts": "Investment Accounts",
        "retirement_accounts": "Retirement Accounts",
        "real_estate_value": "Real Estate",
        "vehicle_value": "Vehicle(s)",
        "other_assets": "Other Assets",
    }
    for key, label in asset_types.items():
        if assets[key] > 0:
            details.append({
                "category": label,
                "estimated_value_aed": str(assets[key]),
                "valuation_date": (date.today() - timedelta(days=rng.randint(0, 180))).isoformat(),
            })
    return details


def _build_liability_details(liabilities: dict[str, Decimal], rng: random.Random) -> list[dict[str, Any]]:
    """Build detailed liability records for the liability_details JSONB field."""
    details = []
    liability_types = {
        "mortgage_balance": "Mortgage",
        "personal_loans": "Personal Loans",
        "credit_card_debt": "Credit Card Debt",
        "student_loans": "Student Loans",
        "other_liabilities": "Other Liabilities",
    }
    for key, label in liability_types.items():
        if liabilities[key] > 0:
            details.append({
                "category": label,
                "outstanding_balance_aed": str(liabilities[key]),
                "monthly_payment_aed": str((liabilities[key] * Decimal(str(rng.uniform(0.01, 0.05)))).quantize(Decimal("0.01"))),
            })
    return details


def _build_income_sources(profile: ApplicantProfile, rng: random.Random) -> list[dict[str, Any]]:
    """Build income source records for the income_sources JSONB field."""
    sources = []
    if profile.monthly_salary > 0:
        sources.append({
            "source": "Employment Salary",
            "amount_aed": str(profile.monthly_salary),
            "frequency": "Monthly",
        })
    if profile.other_income > 0:
        sources.append({
            "source": "Other Income",
            "amount_aed": str(profile.other_income),
            "frequency": "Monthly",
        })
    if profile.housing_status == "owned" and rng.random() < 0.3:
        rent_income = Decimal(str(rng.uniform(1000, 8000))).quantize(Decimal("0.01"))
        sources.append({
            "source": "Rental Income",
            "amount_aed": str(rent_income),
            "frequency": "Monthly",
        })
    return sources


def _generate_xlsx(data: dict[str, Any], output_dir: Path, seed: int) -> Path:
    """Generate a formatted XLSX financial statement from assets/liabilities data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets & Liabilities"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    section_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = {"style": "thin"}

    row = 1
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = f"Assets & Liabilities Statement"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = f"Applicant: {data['applicant_name']}"
    ws[f"A{row}"].font = Font(size=11)
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = f"Statement Date: {data['statement_date']}"
    ws[f"A{row}"].font = Font(size=11)
    row += 2

    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "ASSETS"
    cell.fill = section_fill
    cell.font = section_font
    row += 1

    ws[f"A{row}"].value = "Category"
    ws[f"B{row}"].value = "Value (AED)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.font = header_font
    row += 1

    asset_rows = [
        ("Cash & Deposits", data.get("cash_and_deposits", Decimal("0"))),
        ("Savings Accounts", data.get("savings_accounts", Decimal("0"))),
        ("Investment Accounts", data.get("investment_accounts", Decimal("0"))),
        ("Retirement Accounts", data.get("retirement_accounts", Decimal("0"))),
        ("Real Estate", data.get("real_estate_value", Decimal("0"))),
        ("Vehicle(s)", data.get("vehicle_value", Decimal("0"))),
        ("Other Assets", data.get("other_assets", Decimal("0"))),
    ]
    for label, value in asset_rows:
        ws[f"A{row}"].value = label
        ws[f"B{row}"].value = float(value)
        ws[f"B{row}"].number_format = '#,##0.00'
        row += 1

    ws[f"A{row}"].value = "Total Assets"
    ws[f"B{row}"].value = float(data["total_assets"])
    ws[f"B{row}"].number_format = '#,##0.00'
    ws[f"A{row}"].fill = total_fill
    ws[f"A{row}"].font = total_font
    ws[f"B{row}"].fill = total_fill
    ws[f"B{row}"].font = total_font
    row += 2

    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "LIABILITIES"
    cell.fill = section_fill
    cell.font = section_font
    row += 1

    ws[f"A{row}"].value = "Category"
    ws[f"B{row}"].value = "Balance (AED)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.font = header_font
    row += 1

    liability_rows = [
        ("Mortgage Balance", data.get("mortgage_balance", Decimal("0"))),
        ("Personal Loans", data.get("personal_loans", Decimal("0"))),
        ("Credit Card Debt", data.get("credit_card_debt", Decimal("0"))),
        ("Student Loans", data.get("student_loans", Decimal("0"))),
        ("Other Liabilities", data.get("other_liabilities", Decimal("0"))),
    ]
    for label, value in liability_rows:
        ws[f"A{row}"].value = label
        ws[f"B{row}"].value = float(value)
        ws[f"B{row}"].number_format = '#,##0.00'
        row += 1

    ws[f"A{row}"].value = "Total Liabilities"
    ws[f"B{row}"].value = float(data["total_liabilities"])
    ws[f"B{row}"].number_format = '#,##0.00'
    ws[f"A{row}"].fill = total_fill
    ws[f"A{row}"].font = total_font
    ws[f"B{row}"].fill = total_fill
    ws[f"B{row}"].font = total_font
    row += 2

    ws[f"A{row}"].value = "NET WORTH"
    ws[f"B{row}"].value = float(data["net_worth"])
    ws[f"B{row}"].number_format = '#,##0.00'
    ws[f"A{row}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws[f"A{row}"].font = Font(bold=True, size=12)
    ws[f"B{row}"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws[f"B{row}"].font = Font(bold=True, size=12)
    row += 2

    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value = "INCOME"
    cell.fill = section_fill
    cell.font = section_font
    row += 1

    ws[f"A{row}"].value = "Monthly Income"
    ws[f"B{row}"].value = float(data.get("monthly_income", Decimal("0")))
    ws[f"B{row}"].number_format = '#,##0.00'
    row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    output_path = output_dir / f"assets_liabilities_{seed}.xlsx"
    wb.save(str(output_path))
    return output_path


def generate_assets_liabilities(
    profile: ApplicantProfile,
    credit_data: dict | None,
    seed: int,
    credit_facilities: list[dict] | None = None,
) -> tuple[dict[str, Any], Path]:
    """Generate assets/liabilities structured data and XLSX file.

    Args:
        profile: ApplicantProfile seed object for cross-document consistency.
        credit_data: Credit report data dict (from credit_report_generator).
                     Used for credit_card_debt cross-document consistency.
        seed: Random seed for reproducibility.
        credit_facilities: Optional list of credit facility records for coordination.

    Returns:
        Tuple of (assets_liabilities_data dict, XLSX file path).
        The dict maps to the assets_liabilities_data schema (18 fields).
    """
    rng = random.Random(seed)

    assets = _generate_asset_categories(profile, rng)
    liabilities = _generate_liability_categories(profile, credit_data, credit_facilities, rng)

    total_assets = sum(assets.values(), Decimal("0.00"))
    total_liabilities = sum(liabilities.values(), Decimal("0.00"))
    net_worth = total_assets - total_liabilities

    statement_date = date.today() - timedelta(days=rng.randint(0, 180))

    data: dict[str, Any] = {
        "applicant_name": profile.full_name_en,
        "statement_date": statement_date.isoformat(),
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "net_worth": net_worth,
        "cash_and_deposits": assets["cash_and_deposits"],
        "savings_accounts": assets["savings_accounts"],
        "investment_accounts": assets["investment_accounts"],
        "retirement_accounts": assets["retirement_accounts"],
        "real_estate_value": assets["real_estate_value"],
        "vehicle_value": assets["vehicle_value"],
        "other_assets": assets["other_assets"],
        "mortgage_balance": liabilities["mortgage_balance"],
        "personal_loans": liabilities["personal_loans"],
        "credit_card_debt": liabilities["credit_card_debt"],
        "student_loans": liabilities["student_loans"],
        "other_liabilities": liabilities["other_liabilities"],
        "monthly_income": profile.total_monthly_income,
        "income_sources": _build_income_sources(profile, rng),
        "asset_details": _build_asset_details(assets, rng),
        "liability_details": _build_liability_details(liabilities, rng),
    }

    output_dir = Path(__file__).parent.parent.parent / "output" / f"applicant_{seed}"
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = _generate_xlsx(data, output_dir, seed)

    return data, xlsx_path
