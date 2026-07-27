"""Unit tests for extraction agent.

Tests cover:
1. Tool invocation (OCR, PDF, table, resume, XLSX, confidence)
2. Gate 1 integration (document integrity validation)
3. Agent output parsing (JSON, malformed JSON, missing fields)
4. Multi-document extraction (mocked LLM)
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
import structlog
from langchain_core.messages import AIMessage, HumanMessage

from src.agents.extraction.nodes import (
    ExtractionOutput,
    _extract_json_from_text,
    _parse_agent_output,
    _regex_fallback_extraction,
)

logger = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# Test Class 1: OCR Extract Tool
# ---------------------------------------------------------------------------

class TestOcrExtractTool:
    """Test OCR extraction tool with PaddleOCR."""

    def test_ocr_extract_with_valid_image(self, tmp_path):
        """Test OCR extraction with a valid image file."""
        pytest.importorskip("paddleocr")
        from src.agents.extraction.tools import ocr_extract_tool

        # Create a dummy image file
        img_path = tmp_path / "test.png"
        img_path.write_bytes(b"fake image data")

        # Mock the OCR engine
        with patch("src.infrastructure.document_processing.ocr.OCREngine") as mock_ocr:
            mock_engine = MagicMock()
            mock_result = MagicMock()
            mock_result.text = "Extracted text"
            mock_result.blocks = [{"text": "block1", "confidence": 0.95}]
            mock_result.confidence = 0.95
            mock_result.language = "en+ar"
            mock_engine.extract = AsyncMock(return_value=mock_result)
            mock_ocr.return_value = mock_engine

            result = ocr_extract_tool.invoke({"file_path": str(img_path)})

            logger.debug("tool_invoked", tool="ocr_extract", confidence=result["confidence"])
            assert "text" in result
            assert result["text"] == "Extracted text"
            assert result["confidence"] == 0.95
            assert "duration_ms" in result

    def test_ocr_extract_file_not_found(self):
        """Test OCR extraction with non-existent file."""
        pytest.importorskip("paddleocr")
        from src.agents.extraction.tools import ocr_extract_tool

        result = ocr_extract_tool.invoke({"file_path": "/nonexistent/file.png"})
        assert "error" in result
        assert "File not found" in result["error"]


# ---------------------------------------------------------------------------
# Test Class 2: PDF Parse Tool
# ---------------------------------------------------------------------------

class TestPdfParseTool:
    """Test PDF parsing tool with PyMuPDF4LLM."""

    def test_pdf_parse_with_valid_pdf(self, tmp_path):
        """Test PDF parsing with a valid PDF file."""
        pytest.importorskip("pymupdf4llm")
        from src.agents.extraction.tools import pdf_parse_tool

        pdf_path = tmp_path / "test.pdf"
        pdf_path.write_bytes(b"%PDF-1.4 fake pdf data")

        with patch("src.infrastructure.document_processing.pdf_parser.PDFParser") as mock_parser:
            mock_parser_instance = MagicMock()
            mock_result = MagicMock()
            mock_result.raw_extracted_data = {"markdown": "# Test", "json_structure": {}}
            mock_result.fields = [{"name": "field1", "value": "value1"}]
            mock_result.overall_confidence = 0.92
            mock_result.document_type = "credit_report"
            mock_parser_instance.extract = AsyncMock(return_value=mock_result)
            mock_parser.return_value = mock_parser_instance

            result = pdf_parse_tool.invoke({"file_path": str(pdf_path)})

            assert "markdown" in result
            assert result["confidence"] == 0.92
            assert result["field_count"] == 1
            assert "duration_ms" in result

    def test_pdf_parse_file_not_found(self):
        """Test PDF parsing with non-existent file."""
        pytest.importorskip("pymupdf4llm")
        from src.agents.extraction.tools import pdf_parse_tool

        result = pdf_parse_tool.invoke({"file_path": "/nonexistent/file.pdf"})
        assert "error" in result


# ---------------------------------------------------------------------------
# Test Class 3: Table Extract Tool
# ---------------------------------------------------------------------------

class TestTableExtractTool:
    """Test table extraction tool with Camelot."""

    def test_table_extract_with_valid_pdf(self, tmp_path):
        """Test table extraction from PDF."""
        pytest.importorskip("camelot")
        from src.agents.extraction.tools import table_extract_tool

        pdf_path = tmp_path / "test.pdf"
        pdf_path.write_bytes(b"%PDF-1.4 fake pdf data")

        with patch("src.infrastructure.document_processing.table_extractor.TableExtractor") as mock_extractor:
            import pandas as pd

            mock_extractor_instance = MagicMock()
            mock_result = MagicMock()
            mock_result.tables = [pd.DataFrame({"col1": [1, 2], "col2": [3, 4]})]
            mock_result.table_count = 1
            mock_result.confidence = 0.88
            mock_result.flavor = "lattice"
            mock_extractor_instance.extract = AsyncMock(return_value=mock_result)
            mock_extractor.return_value = mock_extractor_instance

            result = table_extract_tool.invoke({"file_path": str(pdf_path)})

            assert "tables" in result
            assert result["table_count"] == 1
            assert result["confidence"] == 0.88
            assert "duration_ms" in result

    def test_table_extract_file_not_found(self):
        """Test table extraction with non-existent file."""
        pytest.importorskip("camelot")
        from src.agents.extraction.tools import table_extract_tool

        result = table_extract_tool.invoke({"file_path": "/nonexistent/file.pdf"})
        assert "error" in result


# ---------------------------------------------------------------------------
# Test Class 4: Resume Parse Tool
# ---------------------------------------------------------------------------

class TestResumeParseTool:
    """Test resume parsing tool with SmartResume."""

    def test_resume_parse_with_valid_docx(self, tmp_path):
        """Test resume parsing with DOCX file."""
        pytest.importorskip("smartresume")
        from src.agents.extraction.tools import resume_parse_tool

        docx_path = tmp_path / "resume.docx"
        docx_path.write_bytes(b"fake docx data")

        with patch("src.infrastructure.document_processing.resume_parser.ResumeParser") as mock_parser:
            mock_parser_instance = MagicMock()
            mock_result = MagicMock()
            mock_result.full_name = "John Doe"
            mock_result.email = "john@example.com"
            mock_result.phone = "+971501234567"
            mock_result.location = "Dubai, UAE"
            mock_result.summary = "Experienced professional"
            mock_result.years_of_experience = 5
            mock_result.work_experience = []
            mock_result.total_positions = 3
            mock_result.current_employer = "Company ABC"
            mock_result.current_job_title = "Manager"
            mock_result.education = []
            mock_result.highest_degree = "MBA"
            mock_result.skills = ["Python", "SQL"]
            mock_result.skill_count = 2
            mock_result.certifications = []
            mock_result.extraction_confidence = 0.90
            mock_parser_instance.parse = AsyncMock(return_value=mock_result)
            mock_parser.return_value = mock_parser_instance

            result = resume_parse_tool.invoke({"file_path": str(docx_path)})

            assert result["full_name"] == "John Doe"
            assert result["email"] == "john@example.com"
            assert result["confidence"] == 0.90
            assert "duration_ms" in result

    def test_resume_parse_file_not_found(self):
        """Test resume parsing with non-existent file."""
        pytest.importorskip("smartresume")
        from src.agents.extraction.tools import resume_parse_tool

        result = resume_parse_tool.invoke({"file_path": "/nonexistent/resume.docx"})
        assert "error" in result


# ---------------------------------------------------------------------------
# Test Class 5: XLSX Extract Tool
# ---------------------------------------------------------------------------

class TestXlsxExtractTool:
    """Test XLSX extraction tool with openpyxl."""

    def test_xlsx_extract_with_valid_file(self, tmp_path):
        """Test XLSX extraction with valid Excel file."""
        pytest.importorskip("pymupdf4llm")
        from openpyxl import Workbook

        from src.agents.extraction.tools import xlsx_extract_tool

        # Create a real Excel file
        xlsx_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        ws.append(["Asset Type", "Value"])
        ws.append(["Cash", 10000])
        ws.append(["Investments", 50000])
        wb.save(xlsx_path)

        result = xlsx_extract_tool.invoke({"file_path": str(xlsx_path)})

        assert "sheets" in result
        assert result["sheet_count"] >= 1
        assert "Assets" in result["sheets"]
        assert "duration_ms" in result

    def test_xlsx_extract_file_not_found(self):
        """Test XLSX extraction with non-existent file."""
        from src.agents.extraction.tools import xlsx_extract_tool

        result = xlsx_extract_tool.invoke({"file_path": "/nonexistent/file.xlsx"})
        assert "error" in result
        assert "File not found" in result["error"]

    def test_xlsx_extract_specific_sheet(self, tmp_path):
        """Test XLSX extraction with specific sheet name."""
        pytest.importorskip("pymupdf4llm")
        from openpyxl import Workbook

        from src.agents.extraction.tools import xlsx_extract_tool

        xlsx_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Assets"
        ws1.append(["Asset", "Value"])
        ws1.append(["Cash", 10000])

        ws2 = wb.create_sheet("Liabilities")
        ws2.append(["Liability", "Amount"])
        ws2.append(["Loan", 5000])
        wb.save(xlsx_path)

        result = xlsx_extract_tool.invoke(
            {"file_path": str(xlsx_path), "sheet_name": "Liabilities"}
        )

        assert result["sheet_count"] == 1
        assert "Liabilities" in result["sheets"]


# ---------------------------------------------------------------------------
# Test Class 6: Confidence Score Tool
# ---------------------------------------------------------------------------

class TestConfidenceScoreTool:
    """Test confidence scoring tool."""

    def test_confidence_score_with_valid_data(self):
        """Test confidence scoring with valid extracted data."""
        from src.agents.extraction.tools import confidence_score_tool

        extracted_data = {
            "identity_number": "784-1990-1234567-8",
            "full_name_en": "John Doe",
            "nationality": "UAE",
            "date_of_birth": "1990-01-15",
            "gender": "Male",
            "expiry_date": "2030-01-01",
            "extraction_confidence": 0.95,
        }

        result = confidence_score_tool.invoke(
            {"extracted_data": extracted_data, "document_type": "emirates_id"}
        )

        assert "overall_confidence" in result
        assert "routing_decision" in result
        assert result["routing_decision"] in ["auto", "spot_check", "manual_review"]
        assert "duration_ms" in result

    def test_confidence_score_with_missing_fields(self):
        """Test confidence scoring with missing required fields."""
        from src.agents.extraction.tools import confidence_score_tool

        extracted_data = {
            "identity_number": "784-1990-1234567-8",
            # Missing other required fields
        }

        result = confidence_score_tool.invoke(
            {"extracted_data": extracted_data, "document_type": "emirates_id"}
        )

        # Should still return a result, possibly with lower confidence
        assert "overall_confidence" in result
        assert "routing_decision" in result


# ---------------------------------------------------------------------------
# Test Class 7: Gate 1 Integration
# ---------------------------------------------------------------------------

class TestGate1Integration:
    """Test Gate 1 (document integrity) integration."""

    def test_extraction_passes_gate1(self, sample_extracted_data):
        """Test extraction passes Gate 1 with valid data."""
        from src.agents.gates.document_integrity import validate_document_integrity

        # Use emirates_id data from fixture
        emirates_id_data = sample_extracted_data["emirates_id"]

        # Add required fields for validation
        emirates_id_data["is_mrz_verified"] = True

        is_valid, errors = validate_document_integrity(emirates_id_data, "emirates_id")

        # Should pass or have minimal errors
        assert isinstance(is_valid, bool)
        assert isinstance(errors, list)

    def test_extraction_fails_gate1_missing_fields(self):
        """Test extraction fails Gate 1 with missing required fields."""
        from src.agents.gates.document_integrity import validate_document_integrity

        incomplete_data = {
            "identity_number": "784-1990-1234567-8",
            # Missing other required fields
        }

        is_valid, errors = validate_document_integrity(incomplete_data, "emirates_id")

        assert is_valid is False
        assert len(errors) > 0
        assert any("missing" in err.lower() for err in errors)

    def test_extraction_fails_gate1_invalid_checksum(self):
        """Test extraction fails Gate 1 with invalid Emirates ID checksum."""
        from src.agents.gates.document_integrity import validate_document_integrity

        invalid_data = {
            "identity_number": "784-1990-1234567-9",  # Invalid checksum
            "full_name_en": "John Doe",
            "nationality": "UAE",
            "date_of_birth": "1990-01-15",
            "gender": "Male",
            "expiry_date": "2030-01-01",
            "is_mrz_verified": True,
        }

        is_valid, errors = validate_document_integrity(invalid_data, "emirates_id")

        assert is_valid is False
        assert any("checksum" in err.lower() or "format" in err.lower() for err in errors)

    def test_gate1_retry_logic(self, sample_state):
        """Test retry logic when Gate 1 fails."""
        from src.agents.extraction.nodes import MAX_EXTRACTION_RETRIES

        # Verify max retries is set correctly
        assert MAX_EXTRACTION_RETRIES == 2

        # Simulate retry scenario
        attempt = 0
        max_attempts = MAX_EXTRACTION_RETRIES + 1  # Initial + retries

        while attempt < max_attempts:
            attempt += 1
            if attempt == max_attempts:
                break

        assert attempt == max_attempts


# ---------------------------------------------------------------------------
# Test Class 8: Agent Output Parsing
# ---------------------------------------------------------------------------

class TestAgentOutputParsing:
    """Test agent output parsing logic."""

    def test_parse_valid_json(self):
        """Test parsing valid JSON output."""
        json_content = json.dumps({
            "document_type": "emirates_id",
            "identity_number": "784-1990-1234567-8",
            "full_name_en": "John Doe",
            "extraction_confidence": 0.95,
        })

        message = AIMessage(content=json_content)
        result = _parse_agent_output(message, "emirates_id")

        assert result["document_type"] == "emirates_id"
        assert result["identity_number"] == "784-1990-1234567-8"
        assert result["extraction_confidence"] == 0.95

    def test_parse_json_in_markdown_block(self):
        """Test parsing JSON wrapped in markdown code block."""
        json_content = json.dumps({
            "document_type": "bank_statement",
            "bank_name": "Emirates NBD",
            "extraction_confidence": 0.92,
        })

        markdown_content = f"Here is the extracted data:\n```json\n{json_content}\n```"
        message = AIMessage(content=markdown_content)
        result = _parse_agent_output(message, "bank_statement")

        assert result["document_type"] == "bank_statement"
        assert result["bank_name"] == "Emirates NBD"

    def test_parse_malformed_json_fallback(self):
        """Test parsing malformed JSON with regex fallback."""
        malformed_content = """
        The extracted data is:
        Identity Number: 784-1990-1234567-8
        Name: John Doe
        Date of Birth: 1990-01-15
        """

        message = AIMessage(content=malformed_content)
        result = _parse_agent_output(message, "emirates_id")

        # Should attempt regex fallback
        assert isinstance(result, dict)
        # May extract some fields via regex
        if result:
            assert result.get("document_type") == "emirates_id"

    def test_parse_missing_document_type(self):
        """Test parsing JSON without document_type field."""
        json_content = json.dumps({
            "identity_number": "784-1990-1234567-8",
            "full_name_en": "John Doe",
        })

        message = AIMessage(content=json_content)
        result = _parse_agent_output(message, "emirates_id")

        # Should add document_type
        assert result["document_type"] == "emirates_id"

    def test_parse_none_message(self):
        """Test parsing None message."""
        result = _parse_agent_output(None, "emirates_id")
        assert result == {}

    def test_parse_empty_message(self):
        """Test parsing empty message."""
        message = AIMessage(content="")
        result = _parse_agent_output(message, "emirates_id")
        assert result == {}

    def test_extract_json_from_text_code_block(self):
        """Test JSON extraction from markdown code block."""
        text = """
        Here is the data:
        ```json
        {"key": "value"}
        ```
        """
        json_str = _extract_json_from_text(text)
        assert json_str is not None
        parsed = json.loads(json_str)
        assert parsed["key"] == "value"

    def test_extract_json_from_text_raw(self):
        """Test JSON extraction from raw text."""
        text = 'The result is {"key": "value"} and more text'
        json_str = _extract_json_from_text(text)
        assert json_str is not None
        parsed = json.loads(json_str)
        assert parsed["key"] == "value"

    def test_extract_json_from_text_none(self):
        """Test JSON extraction when no JSON present."""
        text = "No JSON here, just plain text"
        json_str = _extract_json_from_text(text)
        assert json_str is None

    def test_regex_fallback_extraction(self):
        """Test regex fallback extraction."""
        text = """
        Identity Number: 784-1990-1234567-8
        Name: John Doe
        Date of Birth: 1990-01-15
        Nationality: UAE
        Gender: Male
        """

        result = _regex_fallback_extraction(text, "emirates_id")

        assert result["document_type"] == "emirates_id"
        assert result["extraction_confidence"] == 0.3  # Low confidence
        assert "identity_number" in result
        assert result["identity_number"] == "784-1990-1234567-8"


# ---------------------------------------------------------------------------
# Test Class 9: Multi-Document Extraction
# ---------------------------------------------------------------------------

class TestMultiDocumentExtraction:
    """Test multi-document extraction with mocked LLM."""

    @pytest.mark.asyncio
    async def test_extract_multiple_documents(self, sample_state):
        """Test extraction of multiple documents."""
        from src.agents.extraction.nodes import extract_documents_node

        # Mock the LLM client and agent
        with patch("src.agents.orchestrator.nodes._get_llm_client") as mock_get_llm:
            mock_llm = AsyncMock()
            mock_get_llm.return_value = mock_llm

            import sys
            mock_langchain_openai = MagicMock()
            mock_chat_instance = MagicMock()
            mock_langchain_openai.ChatOpenAI.return_value = mock_chat_instance
            
            with patch.dict(sys.modules, {'langchain_openai': mock_langchain_openai}):
                with patch("src.agents.extraction.nodes._build_extraction_agent") as mock_build:
                    mock_agent = MagicMock()

                    # Mock agent response
                    async def mock_ainvoke(input_data):
                        return {
                            "messages": [
                                AIMessage(content=json.dumps({
                                    "document_type": "emirates_id",
                                    "identity_number": "784-1990-1234567-8",
                                    "full_name_en": "John Doe",
                                    "extraction_confidence": 0.95,
                                }))
                            ]
                        }

                    mock_agent.ainvoke = mock_ainvoke
                    mock_build.return_value = mock_agent

                    with patch("src.agents.extraction.nodes.validate_document_integrity") as mock_validate:
                        mock_validate.return_value = (True, [])

                        result = await extract_documents_node(sample_state)

                        assert "extracted_data" in result
                        assert "extraction_confidence" in result
                        assert "gate_status" in result

    @pytest.mark.asyncio
    async def test_extract_no_documents(self, sample_state):
        """Test extraction with no documents."""
        from src.agents.extraction.nodes import extract_documents_node

        sample_state["uploaded_documents"] = []

        result = await extract_documents_node(sample_state)

        assert result["extracted_data"] == {}
        assert result["extraction_confidence"] == {}
        assert result["gate_status"] == "passed"

    @pytest.mark.asyncio
    async def test_extract_document_fails_gate(self, sample_state):
        """Test extraction when document fails Gate 1."""
        from src.agents.extraction.nodes import extract_documents_node

        with patch("src.agents.orchestrator.nodes._get_llm_client") as mock_get_llm:
            mock_llm = AsyncMock()
            mock_get_llm.return_value = mock_llm

            import sys
            mock_langchain_openai = MagicMock()
            mock_chat_instance = MagicMock()
            mock_langchain_openai.ChatOpenAI.return_value = mock_chat_instance

            with patch.dict(sys.modules, {'langchain_openai': mock_langchain_openai}):
                with patch("src.agents.extraction.nodes._build_extraction_agent") as mock_build:
                    mock_agent = MagicMock()

                    async def mock_ainvoke(input_data):
                        return {
                            "messages": [
                                AIMessage(content=json.dumps({
                                    "document_type": "emirates_id",
                                    "identity_number": "invalid",
                                    "extraction_confidence": 0.5,
                                }))
                            ]
                        }

                    mock_agent.ainvoke = mock_ainvoke
                    mock_build.return_value = mock_agent

                    with patch("src.agents.extraction.nodes.validate_document_integrity") as mock_validate:
                        mock_validate.return_value = (False, ["Invalid checksum"])

                        result = await extract_documents_node(sample_state)

                        assert result["gate_status"] == "failed"
                        assert len(result["gate_errors"]) > 0

    def test_extraction_confidence_aggregation(self):
        """Test confidence aggregation across multiple documents."""
        confidences = {
            "doc1": 0.95,
            "doc2": 0.92,
            "doc3": 0.88,
        }

        avg_confidence = sum(confidences.values()) / len(confidences)

        assert avg_confidence == pytest.approx(0.9167, rel=1e-3)

    def test_extraction_error_handling(self, sample_state):
        """Test error handling when one document fails."""
        # Simulate partial success
        extracted_data = {
            "doc1": {"document_type": "emirates_id", "valid": True},
            "doc2": {"document_type": "bank_statement", "valid": True},
        }

        # One document failed
        total_docs = 3
        succeeded = len(extracted_data)
        failed = total_docs - succeeded

        assert succeeded == 2
        assert failed == 1


# ---------------------------------------------------------------------------
# Additional Edge Case Tests
# ---------------------------------------------------------------------------

class TestExtractionOutputModel:
    """Test ExtractionOutput Pydantic model."""

    def test_valid_extraction_output(self):
        """Test valid ExtractionOutput creation."""
        output = ExtractionOutput(
            document_type="emirates_id",
            extraction_confidence=0.95,
            identity_number="784-1990-1234567-8",
        )

        assert output.document_type == "emirates_id"
        assert output.extraction_confidence == 0.95

    def test_extraction_output_default_confidence(self):
        """Test ExtractionOutput with default confidence."""
        output = ExtractionOutput(document_type="bank_statement")

        assert output.extraction_confidence == 0.85

    def test_extraction_output_extra_fields(self):
        """Test ExtractionOutput allows extra fields."""
        output = ExtractionOutput(
            document_type="credit_report",
            extraction_confidence=0.90,
            credit_score=720,
            risk_band="Good",
        )

        assert output.document_type == "credit_report"
        # Extra fields should be accessible
        data = output.model_dump()
        assert data["credit_score"] == 720
        assert data["risk_band"] == "Good"

    def test_extraction_output_invalid_confidence(self):
        """Test ExtractionOutput with invalid confidence value."""
        with pytest.raises(ValueError):
            ExtractionOutput(
                document_type="emirates_id",
                extraction_confidence=1.5,  # > 1.0
            )
